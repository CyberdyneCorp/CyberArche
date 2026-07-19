"""FileExtractorPort adapter: PDF (pypdf), Excel (openpyxl), CSV, Markdown,
Word (python-docx), text.

Tabular sources become `table` blocks (rows/columns match the sheet); Markdown
and Word become structured blocks (headings, lists, quotes, code, tables);
plain text becomes paragraph blocks (ai-agent + block-editor + document-import
specs).
"""

from __future__ import annotations

import csv
import io
import uuid
import zipfile
from pathlib import PurePosixPath

from docx import Document as Docx
from docx.opc.exceptions import PackageNotFoundError
from docx.table import Table as DocxTable
from openpyxl import load_workbook
from pypdf import PdfReader

from cyberarche.domain.errors import ValidationFailed
from cyberarche.domain.markdown_blocks import markdown_to_blocks


def _new_id() -> str:
    return uuid.uuid4().hex


def _block(block_type: str, data: dict) -> dict:
    return {"id": _new_id(), "type": block_type, "data": data}


# python-docx paragraph style name -> heading level.
_HEADING_LEVELS = {"Heading 1": 1, "Heading 2": 2, "Heading 3": 3}


def _table_block(rows: list[list[str]], *, source: str) -> dict:
    header, body = (rows[0], rows[1:]) if rows else ([], [])
    return _block(
        "table",
        {"header": header, "rows": body, "source": source},
    )


def _paragraphs(text: str) -> list[dict]:
    chunks = [chunk.strip() for chunk in text.split("\n\n")]
    return [_block("paragraph", {"text": chunk}) for chunk in chunks if chunk]


class FileExtractor:
    def extract_blocks(self, *, filename: str, content: bytes) -> list[dict]:
        suffix = PurePosixPath(filename.lower()).suffix
        if suffix == ".pdf":
            return self._pdf(content)
        if suffix == ".csv":
            return self._csv(content, filename)
        if suffix in (".xlsx", ".xlsm"):
            return self._excel(content, filename)
        if suffix in (".md", ".markdown"):
            text = content.decode("utf-8", errors="replace")
            return markdown_to_blocks(text, _new_id)
        if suffix == ".docx":
            return self._docx(content)
        if suffix in (".txt", ""):
            return _paragraphs(content.decode("utf-8", errors="replace"))
        raise ValidationFailed(f"unsupported file type: {suffix}")

    def extract_table(
        self, *, filename: str, content: bytes
    ) -> tuple[list[str], list[list[str]]]:
        """Parse a CSV or Excel sheet into ``(header, rows)`` — the first row is
        the header, the rest are data rows. Empty input yields ``([], [])``.
        Reuses the same csv/openpyxl parsing as `extract_blocks`."""
        suffix = PurePosixPath(filename.lower()).suffix
        if suffix == ".csv":
            rows = _csv_rows(content)
        elif suffix in (".xlsx", ".xlsm"):
            rows = _first_sheet_rows(content)
        else:
            raise ValidationFailed(f"not a spreadsheet file: {suffix}")
        return (rows[0], rows[1:]) if rows else ([], [])

    def _pdf(self, content: bytes) -> list[dict]:
        reader = PdfReader(io.BytesIO(content))
        blocks: list[dict] = []
        for page in reader.pages:
            blocks.extend(_paragraphs(page.extract_text() or ""))
        return blocks

    def _csv(self, content: bytes, filename: str) -> list[dict]:
        rows = _csv_rows(content)
        if not rows:
            return []
        return [_table_block(rows, source=filename)]

    def _docx(self, content: bytes) -> list[dict]:
        try:
            document = Docx(io.BytesIO(content))
        except (PackageNotFoundError, zipfile.BadZipFile, ValueError, KeyError) as exc:
            raise ValidationFailed("could not read Word document") from exc
        blocks: list[dict] = []
        for item in document.iter_inner_content():
            if isinstance(item, DocxTable):
                blocks.append(_docx_table_block(item))
                continue
            block = _docx_paragraph_block(item)
            if block is not None:
                blocks.append(block)
        return blocks

    def _excel(self, content: bytes, filename: str) -> list[dict]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        blocks: list[dict] = []
        for sheet in workbook.worksheets:
            rows = _sheet_rows(sheet)
            if not rows:
                continue
            blocks.append(_block("heading", {"text": sheet.title, "level": 2}))
            blocks.append(_table_block(rows, source=f"{filename}#{sheet.title}"))
        return blocks


def _docx_paragraph_block(paragraph) -> dict | None:
    """Map a Word paragraph to a block by its style, skipping empty paragraphs."""
    text = paragraph.text.strip()
    if not text:
        return None
    style = paragraph.style.name if paragraph.style else ""
    level = _HEADING_LEVELS.get(style)
    if level is not None:
        return _block("heading", {"text": text, "level": level})
    if style.startswith("List Bullet"):
        return _block("bulleted_list", {"text": text})
    if style.startswith("List Number"):
        return _block("numbered_list", {"text": text})
    if style == "Quote":
        return _block("quote", {"text": text})
    return _block("paragraph", {"text": text})


def _docx_table_block(table: DocxTable) -> dict:
    rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
    return _table_block(rows, source="docx")


def _sheet_rows(sheet) -> list[list[str]]:
    return [
        ["" if cell is None else str(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
        if any(cell is not None for cell in row)
    ]


def _csv_rows(content: bytes) -> list[list[str]]:
    """CSV bytes -> non-empty rows (utf-8 with an optional BOM)."""
    text = content.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text)) if row]


def _first_sheet_rows(content: bytes) -> list[list[str]]:
    """The non-empty rows of an Excel workbook's first sheet ([] when empty)."""
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = workbook.worksheets
    return _sheet_rows(sheets[0]) if sheets else []
